import json
import logging
from io import BytesIO
from pathlib import Path
from datetime import datetime

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from app.models import GenerateRequest, ChatRequest, BOPData, UnifiedChatRequest, UnifiedChatResponse
from app.llm_service import generate_bop_from_text, modify_bop, unified_chat, get_resource_size
from app.tools.router import router as tools_router
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 로그 디렉토리 생성
LOG_DIR = Path(__file__).resolve().parent.parent / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

# 로그 파일 경로 (날짜별로 분리)
log_filename = LOG_DIR / f"backend_{datetime.now().strftime('%Y%m%d')}.log"

# 로그 포맷터
log_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")

# 파일 핸들러
file_handler = logging.FileHandler(log_filename, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(log_formatter)

# 콘솔 핸들러
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)  # 콘솔은 INFO 레벨만
console_handler.setFormatter(log_formatter)

# 루트 로거 설정
root_logger = logging.getLogger()
root_logger.setLevel(logging.DEBUG)
root_logger.addHandler(file_handler)
root_logger.addHandler(console_handler)

# HTTP 관련 로거 레벨 조정 (API 키 노출 방지)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("openai").setLevel(logging.WARNING)


# API 키 마스킹 필터 (안전망 - 모든 로그에서 키 패턴을 자동 마스킹)
class APIKeyMaskingFilter(logging.Filter):
    """로그 메시지에서 API 키 패턴을 마스킹합니다."""
    import re as _re
    _patterns = [
        _re.compile(r'(AIza[A-Za-z0-9_-]{30,})'),           # Gemini/Google API 키
        _re.compile(r'(sk-[A-Za-z0-9]{20,})'),              # OpenAI API 키
        _re.compile(r'([?&]key=)([A-Za-z0-9_-]{20,})'),     # URL 쿼리의 key 파라미터
        _re.compile(r'(Bearer\s+)([A-Za-z0-9_-]{20,})'),    # Bearer 토큰
    ]

    def filter(self, record):
        msg = record.getMessage()
        for pattern in self._patterns:
            if pattern.search(msg):
                for p in self._patterns:
                    msg = p.sub(lambda m: m.group(0)[:8] + '***MASKED***' if m.lastindex is None or m.lastindex < 2
                                else m.group(1) + '***MASKED***', msg)
                record.msg = msg
                record.args = None
                break
        return True


# 모든 핸들러에 마스킹 필터 적용
_masking_filter = APIKeyMaskingFilter()
file_handler.addFilter(_masking_filter)
console_handler.addFilter(_masking_filter)

# 시작 메시지
logging.info("=" * 80)
logging.info(f"Backend started - Log file: {log_filename}")
logging.info("=" * 80)

app = FastAPI(title="Backend API", version="1.0.0")

# 전역 예외 핸들러 (500 에러 로깅)
@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    import traceback
    from fastapi.responses import JSONResponse

    # 상세 에러 로깅
    logging.error("=" * 80)
    logging.error("[GLOBAL ERROR] 500 Internal Server Error")
    logging.error(f"[GLOBAL ERROR] Request: {request.method} {request.url}")
    logging.error(f"[GLOBAL ERROR] Exception Type: {type(exc).__name__}")
    logging.error(f"[GLOBAL ERROR] Exception Message: {str(exc)}")
    logging.error(f"[GLOBAL ERROR] Traceback:\n{traceback.format_exc()}")
    logging.error("=" * 80)

    return JSONResponse(
        status_code=500,
        content={"detail": f"Internal server error: {str(exc)}"}
    )

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174", "http://localhost:5175"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(tools_router)


@app.get("/")
async def root():
    return {"message": "Backend Ready"}


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/api/models")
async def get_supported_models():
    """지원하는 LLM 모델 목록 반환"""
    from app.llm import get_supported_models
    return get_supported_models()


@app.post("/api/generate")
async def generate_bop(req: GenerateRequest) -> BOPData:
    """
    사용자 입력을 받아 Gemini API를 통해 BOP를 생성합니다.
    """
    try:
        # LLM 서비스를 통해 BOP 생성
        bop_dict = await generate_bop_from_text(req.user_input)

        # Pydantic 모델로 validation
        bop_data = BOPData(**bop_dict)

        return bop_data

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"BOP 생성 실패: {str(e)}")


@app.post("/api/chat")
async def chat(req: ChatRequest) -> BOPData:
    """
    현재 BOP와 사용자 메시지를 받아 수정된 BOP를 반환합니다.
    """
    try:
        # current_bop을 dict로 변환
        current_bop_dict = req.current_bop.model_dump()

        # LLM 서비스를 통해 BOP 수정
        updated_bop_dict = await modify_bop(current_bop_dict, req.message)

        # Pydantic 모델로 validation
        updated_bop = BOPData(**updated_bop_dict)

        return updated_bop

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"BOP 수정 실패: {str(e)}")


@app.post("/api/chat/unified")
async def unified_chat_endpoint(req: UnifiedChatRequest) -> UnifiedChatResponse:
    """
    통합 채팅 엔드포인트: BOP 생성, 수정, QA를 모두 처리합니다.
    """
    try:
        # current_bop을 dict로 변환 (있는 경우)
        current_bop_dict = req.current_bop.model_dump() if req.current_bop else None

        # LLM 서비스를 통해 통합 처리 (모델, 언어 파라미터 전달)
        response_data = await unified_chat(req.message, current_bop_dict, req.model, req.language)

        # bop_data가 있으면 Pydantic 모델로 validation
        bop_data = None
        if "bop_data" in response_data:
            bop_data = BOPData(**response_data["bop_data"])

        # UnifiedChatResponse 반환
        return UnifiedChatResponse(
            message=response_data["message"],
            bop_data=bop_data
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chat 실패: {str(e)}")


@app.post("/api/export/excel")
async def export_excel(bop: BOPData):
    """
    BOP 데이터를 Excel 파일로 내보냅니다.

    Sheet 1: Overview (Project Info + Resource Masters)
    Sheet 2: BOP Processes
    """
    try:
        wb = Workbook()

        # ==================== Sheet 1: Overview ====================
        ws_overview = wb.active
        ws_overview.title = "Overview"

        # Project Info
        ws_overview.append(["Project Information"])
        ws_overview['A1'].font = Font(bold=True, size=14)
        ws_overview.append(["Project Title", bop.project_title])
        ws_overview.append(["Target UPH", bop.target_uph])
        ws_overview.append([])

        # Equipment Master
        ws_overview.append(["Equipment Master"])
        ws_overview[f'A{ws_overview.max_row}'].font = Font(bold=True, size=12)
        ws_overview.append(["Equipment ID", "Name", "Type"])
        header_row = ws_overview.max_row
        for col in range(1, 4):
            ws_overview.cell(header_row, col).font = Font(bold=True)
            ws_overview.cell(header_row, col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for eq in bop.equipments:
            ws_overview.append([
                eq.equipment_id,
                eq.name,
                eq.type
            ])

        ws_overview.append([])

        # Worker Master
        ws_overview.append(["Worker Master"])
        ws_overview[f'A{ws_overview.max_row}'].font = Font(bold=True, size=12)
        ws_overview.append(["Worker ID", "Name", "Skill Level"])
        header_row = ws_overview.max_row
        for col in range(1, 4):
            ws_overview.cell(header_row, col).font = Font(bold=True)
            ws_overview.cell(header_row, col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for worker in bop.workers:
            ws_overview.append([
                worker.worker_id,
                worker.name,
                worker.skill_level or "-"
            ])

        ws_overview.append([])

        # Material Master
        ws_overview.append(["Material Master"])
        ws_overview[f'A{ws_overview.max_row}'].font = Font(bold=True, size=12)
        ws_overview.append(["Material ID", "Name", "Unit"])
        header_row = ws_overview.max_row
        for col in range(1, 4):
            ws_overview.cell(header_row, col).font = Font(bold=True)
            ws_overview.cell(header_row, col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for material in bop.materials:
            ws_overview.append([
                material.material_id,
                material.name,
                material.unit
            ])

        # ==================== Sheet 2: BOP Processes ====================
        ws_bop = wb.create_sheet("BOP Processes")

        # Header
        ws_bop.append([
            "Process ID", "Parallel Count", "Predecessors", "Successors"
        ])
        header_row = 1
        for col in range(1, 5):
            ws_bop.cell(header_row, col).font = Font(bold=True)
            ws_bop.cell(header_row, col).fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            ws_bop.cell(header_row, col).font = Font(bold=True, color="FFFFFF")
            ws_bop.cell(header_row, col).alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for process in bop.processes:
            parallel_count = len([d for d in bop.process_details if d.process_id == process.process_id])

            predecessor_str = ", ".join(process.predecessor_ids) if process.predecessor_ids else "-"
            successor_str = ", ".join(process.successor_ids) if process.successor_ids else "-"

            ws_bop.append([
                process.process_id,
                parallel_count,
                predecessor_str,
                successor_str
            ])

        # ==================== Sheet 3: Process Details ====================
        ws_details = wb.create_sheet("Process Details")
        ws_details.append([
            "Process ID", "Parallel Index", "Name", "Description",
            "Cycle Time (s)", "Location (X,Y,Z)", "Rotation Y"
        ])
        header_row = 1
        for col in range(1, 8):
            ws_details.cell(header_row, col).font = Font(bold=True)
            ws_details.cell(header_row, col).fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            ws_details.cell(header_row, col).font = Font(bold=True, color="FFFFFF")
            ws_details.cell(header_row, col).alignment = Alignment(horizontal="center", vertical="center")

        for detail in bop.process_details:
            loc = detail.location
            loc_str = f"({loc.x}, {loc.y}, {loc.z})" if loc else "-"
            ws_details.append([
                detail.process_id,
                detail.parallel_index,
                detail.name,
                detail.description or "",
                detail.cycle_time_sec,
                loc_str,
                detail.rotation_y
            ])

        # ==================== Sheet 4: Resource Assignments ====================
        ws_resources = wb.create_sheet("Resource Assignments")
        ws_resources.append([
            "Process ID", "Parallel Index", "Resource Type", "Resource ID",
            "Quantity", "Relative Location (X,Y,Z)", "Rotation Y"
        ])
        header_row = 1
        for col in range(1, 8):
            ws_resources.cell(header_row, col).font = Font(bold=True)
            ws_resources.cell(header_row, col).fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            ws_resources.cell(header_row, col).font = Font(bold=True, color="FFFFFF")
            ws_resources.cell(header_row, col).alignment = Alignment(horizontal="center", vertical="center")

        for ra in bop.resource_assignments:
            rel = ra.relative_location
            rel_str = f"({rel.x}, {rel.y}, {rel.z})" if rel else "-"
            ws_resources.append([
                ra.process_id,
                ra.parallel_index,
                ra.resource_type,
                ra.resource_id,
                ra.quantity,
                rel_str,
                ra.rotation_y
            ])

        # Adjust column widths
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 30
        ws_overview.column_dimensions['C'].width = 20

        ws_bop.column_dimensions['A'].width = 12
        ws_bop.column_dimensions['B'].width = 15
        ws_bop.column_dimensions['C'].width = 20
        ws_bop.column_dimensions['D'].width = 20

        ws_details.column_dimensions['A'].width = 12
        ws_details.column_dimensions['B'].width = 15
        ws_details.column_dimensions['C'].width = 25
        ws_details.column_dimensions['D'].width = 35
        ws_details.column_dimensions['E'].width = 15
        ws_details.column_dimensions['F'].width = 20
        ws_details.column_dimensions['G'].width = 12

        ws_resources.column_dimensions['A'].width = 12
        ws_resources.column_dimensions['B'].width = 15
        ws_resources.column_dimensions['C'].width = 15
        ws_resources.column_dimensions['D'].width = 15
        ws_resources.column_dimensions['E'].width = 10
        ws_resources.column_dimensions['F'].width = 25
        ws_resources.column_dimensions['G'].width = 12

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        safe_filename = "".join(c for c in bop.project_title if c.isalnum() or c in (' ', '_', '-')).strip()
        if not safe_filename:
            safe_filename = "BOP"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={safe_filename}.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export 실패: {str(e)}")


def _get_color_for_equipment_type(equipment_type: str) -> str:
    """Equipment type에 따라 색상 반환"""
    color_map = {
        "robot": "#4a90e2",
        "machine": "#ff6b6b",
        "manual_station": "#50c878"
    }
    return color_map.get(equipment_type, "#888888")


@app.post("/api/export/3d")
async def export_3d(bop: BOPData):
    """
    BOP 데이터를 3D 시각화용 JSON으로 내보냅니다.
    ProcessDetail 위치 + ResourceAssignment 상대 위치로 실제 배치 계산
    """
    try:
        export_data = {
            "project_title": bop.project_title,
            "target_uph": bop.target_uph,
            "processes": [],
            "resources": []
        }

        # Process Details → 3D processes
        for detail in bop.process_details:
            loc = detail.location
            if not loc:
                continue

            process_obj = {
                "process_id": detail.process_id,
                "name": detail.name,
                "parallel_index": detail.parallel_index,
                "position": {
                    "x": loc.x,
                    "y": loc.y,
                    "z": loc.z
                },
                "size": {
                    "width": 4.0,
                    "height": 2.0,
                    "depth": 3.0
                },
                "color": "#e0e0e0"
            }
            export_data["processes"].append(process_obj)

        # Resource Assignments → 3D resources
        # Build detail location map
        detail_loc_map = {}
        for detail in bop.process_details:
            key = (detail.process_id, detail.parallel_index)
            detail_loc_map[key] = detail.location

        for ra in bop.resource_assignments:
            detail_loc = detail_loc_map.get((ra.process_id, ra.parallel_index))
            if not detail_loc:
                continue

            rel = ra.relative_location
            rel_x = rel.x if rel else 0
            rel_y = rel.y if rel else 0
            rel_z = rel.z if rel else 0

            actual_x = detail_loc.x + rel_x
            actual_y = detail_loc.y + rel_y
            actual_z = detail_loc.z + rel_z

            resource_obj = {
                "resource_id": ra.resource_id,
                "resource_type": ra.resource_type,
                "process_id": ra.process_id,
                "parallel_index": ra.parallel_index,
                "quantity": ra.quantity,
                "role": ra.role or "",
                "position": {
                    "x": actual_x,
                    "y": actual_y,
                    "z": actual_z
                }
            }

            if ra.resource_type == "equipment":
                equipment = next((eq for eq in bop.equipments if eq.equipment_id == ra.resource_id), None)
                if equipment:
                    resource_obj["name"] = equipment.name
                    resource_obj["equipment_type"] = equipment.type
                    resource_obj["color"] = _get_color_for_equipment_type(equipment.type)
                    if ra.computed_size:
                        resource_obj["size"] = ra.computed_size.model_dump()
                    else:
                        resource_obj["size"] = get_resource_size("equipment", equipment.type)

            elif ra.resource_type == "worker":
                worker = next((w for w in bop.workers if w.worker_id == ra.resource_id), None)
                if worker:
                    resource_obj["name"] = worker.name
                    resource_obj["color"] = "#50c878"
                    if ra.computed_size:
                        resource_obj["size"] = ra.computed_size.model_dump()
                    else:
                        resource_obj["size"] = get_resource_size("worker")

            elif ra.resource_type == "material":
                material = next((m for m in bop.materials if m.material_id == ra.resource_id), None)
                if material:
                    resource_obj["name"] = material.name
                    resource_obj["unit"] = material.unit
                    resource_obj["color"] = "#ffa500"
                    if ra.computed_size:
                        resource_obj["size"] = ra.computed_size.model_dump()
                    else:
                        resource_obj["size"] = get_resource_size("material")

            export_data["resources"].append(resource_obj)

        # JSON으로 변환
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
        buffer = BytesIO(json_str.encode('utf-8'))
        buffer.seek(0)

        safe_filename = "".join(c for c in bop.project_title if c.isalnum() or c in (' ', '_', '-')).strip()
        if not safe_filename:
            safe_filename = "BOP"

        return StreamingResponse(
            buffer,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={safe_filename}_3d.json"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"3D export 실패: {str(e)}")
